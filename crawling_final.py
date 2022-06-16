#%%
#DART ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
# 종목코드: https://quserinn.github.io/posts/learn-quant-with-python/200522-learn-quant-with-python-2/

from numpy import dtype
import pandas as pd
import openpyxl
import requests
import time
# dart_fss 모듈 이용 전자공시 (DART) 불러오기 참조 페이지
# https://m.blog.naver.com/investgy/222024986530

# 모듈 불러오기
# 모듈 설치시 파이썬 폴더>>pip install dart-fss
#import dart_fss as dart
pd.options.display.float_format = '{:.5f}'.format
ticker="005930"

# api key 변수 설정
# 전자공시 (DART) 페이지에서 OPEN DART API 인증키 신청하기  
#api_key='9e14ce6701fa9dd9e77c0ca7cd437295c8066d06'

#dart.set_api_key(api_key=api_key)

# DART 에 공시된 회사 리스트 불러오기
#corp_list = dart.get_corp_list()
#print(corp_list)
# 삼성전자 검색
#samsung = corp_list.find_by_corp_name('삼성전자', exactly=True)[0]
#samsung = corp_list.find_by_stock_code(f'{ticker}')

# 2012년부터 분기 재무제표 불러오기
#fs1 = samsung.extract_fs(bgn_de='20120101',report_tp='quarter')
#fs2 = samsung.extract_fs(bgn_de='20120101',report_tp='quarter',separate=True)

# 2012년부터 연간 재무제표 불러오기
#fs3 = samsung.extract_fs(bgn_de='20120101')
#fs4 = samsung.extract_fs(bgn_de='20120101',separate=True)

# 재무제표 검색 결과를 엑셀파일로 저장 ( 기본저장위치: 실행폴더/fsdata )
#fs1.save(filename=f"{ticker}_qcon.xlsx",path="./fsdata")
#fs2.save(filename=f"{ticker}_qsepa.xlsx",path="./fsdata")
#fs3.save(filename=f"{ticker}_ycon.xlsx",path="./fsdata")
#fs4.save(filename=f"{ticker}_ysepa.xlsx",path="./fsdata")

# 아버님 원래 템플릿 그릇 
wb = openpyxl.load_workbook("./base.xlsx") 

#sheet1 = openpyxl.load_workbook(f"./fsdata/{ticker}_qcon.xlsx").sheetnames
#sheet2 = openpyxl.load_workbook(f"./fsdata/{ticker}_qsepa.xlsx").sheetnames
#sheet3 = openpyxl.load_workbook(f"./fsdata/{ticker}_ycon.xlsx").sheetnames
#sheet4 = openpyxl.load_workbook(f"./fsdata/{ticker}_ysepa.xlsx").sheetnames

wb.save(filename=f"test_{ticker}.xlsx")

writer = pd.ExcelWriter(f'test_{ticker}.xlsx', mode='a',engine='openpyxl')
'''
count=1
for i in sheet1:
    count+=1
    if count%2==0:
        df2 = pd.read_excel(f"./fsdata/{ticker}_qcon.xlsx", sheet_name=i)
        df2.to_excel(writer,sheet_name=i+"_qcon",float_format='%30.1f')
count=1
for i in sheet2:
    count+=1
    if count%2==0:
        df2 = pd.read_excel(f"./fsdata/{ticker}_qsepa.xlsx", sheet_name=i)
        df2.to_excel(writer,sheet_name=i+"_qsepa",float_format='%30.1f')

count=1
for i in sheet3:
    count+=1
    if count%2==0:
        df2 = pd.read_excel(f"./fsdata/{ticker}_ycon.xlsx", sheet_name=i)
        df2.to_excel(writer,sheet_name=i+"_ycon",float_format='%30.1f')
count=1
for i in sheet4:
    count+=1
    if count%2==0:
        df2 = pd.read_excel(f"./fsdata/{ticker}_ysepa.xlsx", sheet_name=i)
        df2.to_excel(writer,sheet_name=i+"_ysepa",float_format='%30.1f')
'''

# %%
# Financial Summary ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd

url = f'https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={ticker}'
driver = webdriver.Chrome('./chromedriver')
driver.get(url)
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
hd = []
for i in range(33) :
    r = soup.select('tbody')[12]
    r = r.select('tr')[i].select('th')[0].text
    hd.append(r)
headers = []
for i in range(78,86):
    a=str(soup.select('th')[i])
    idx=a.find(">")
    a=a[idx+1:].lstrip()
    headers.append(a[:7])
df = {'주요재무정보': hd}
df = pd.DataFrame(df)
for j in range(len(headers)) :
    fs = []
    for i in range(33) :
        r = soup.select('tbody')[12]
        r = r.select('tr')[i]
        r = r.select('td')[j].text
        fs.append(r)
    df[headers[j]] = fs
    df = pd.DataFrame(df)
print(df)
#df.to_csv("삼성전자.csv", encoding="utf-8-sig")
df.to_excel(writer,sheet_name="Summary",float_format='%30.1f')
driver.close()
#%%
# Financial Summary - 연간■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd

url = f'https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={ticker}'

driver = webdriver.Chrome()
driver.get(url)
time.sleep(5)
driver.find_element_by_id('cns_Tab21').click()
time.sleep(5)
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

hd = []
for i in range(33) :
    r = soup.select('tbody')[12]
    r = r.select('tr')[i].select('th')[0].text
    hd.append(r)
headers = []
for i in range(77,85):
    a=str(soup.select('th')[i])
    idx=a.find(">")
    a=a[idx+1:].lstrip()
    headers.append(a[:7])
    #print(a)
df = {'주요재무정보': hd}
df = pd.DataFrame(df)

for j in range(len(headers)) :
    fs = []
    for i in range(33) :
        #print(j)
        r = soup.select('tbody')[12]
            #print(len(r.select('td')))
        r = r.select('tr')[i]
        r = r.select('td')[j].text
        fs.append(r)

    df[headers[j]] = fs
    df = pd.DataFrame(df)

df.to_excel(writer,sheet_name="Summary_Y",float_format='%30.1f')
driver.close()

#%%
# Financial Summary - 분기■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd

url = f'https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd={ticker}'

driver = webdriver.Chrome()
driver.get(url)
time.sleep(5)
driver.find_element_by_id('cns_Tab22').click()
time.sleep(5)
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')

hd = []
for i in range(33) :
    r = soup.select('tbody')[12]
    r = r.select('tr')[i].select('th')[0].text
    hd.append(r)
headers = []
for i in range(77,85):
    a=str(soup.select('th')[i])
    idx=a.find(">")
    a=a[idx+1:].lstrip()
    headers.append(a[:7])
    #print(a)
df = {'주요재무정보': hd}
df = pd.DataFrame(df)

for j in range(len(headers)) :
    fs = []
    for i in range(33) :
        #print(j)
        r = soup.select('tbody')[12]
            #print(len(r.select('td')))
        r = r.select('tr')[i]
        r = r.select('td')[j].text
        fs.append(r)

    df[headers[j]] = fs
    df = pd.DataFrame(df)
df.to_excel(writer,sheet_name="Summary_S",float_format='%30.1f')
driver.close()
#%%
# 재무분석 - 포괄손익 ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
import time

url = f'https://navercomp.wisereport.co.kr/v2/company/c1030001.aspx?cmp_cd={ticker}'
driver = webdriver.Chrome('./chromedriver')
driver.get(url)
time.sleep(5)
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
hd = []
for i in range(244) :
    r = soup.select('tbody')[5]
    r = r.select('tr')[i].contents[1].text.replace("펼치기","").strip()
    hd.append(r)
headers = []
for i in range(11,18):
    if i != 16 and i != 17:
        ext=str(soup.select('th')[i].select('div')[0])
        idx=ext.find(">")
        ext=ext[idx+1:].lstrip()
        headers.append(ext[:7])
    else:
        ext=str(soup.select('th')[i].select('div')[0])
        idx=ext.find(">")
        ext=ext[idx+1:].lstrip()
        if i == 16:
            headers.append(ext[:10])
        else:
            headers.append(ext[:4])
print(headers)
df = {'주요재무정보': hd}
df = pd.DataFrame(df)
for j in range(0,len(headers)) :
    fs = []
    for i in range(244) :
        r = soup.select('tbody')[5]
        r = r.select('tr')[i] #행 0~
        r = r.select('td')[j+1].text
        fs.append(r)
    df[headers[j]] = fs
df.to_excel(writer,sheet_name="pogual",float_format='%30.1f')
driver.close()
#%%
# 재무분석 - 재무상태표 ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
import time

url = f'https://navercomp.wisereport.co.kr/v2/company/c1030001.aspx?cmp_cd={ticker}'
driver = webdriver.Chrome('./chromedriver')
driver.get(url)
time.sleep(5)
driver.find_element_by_id('rpt_td2').click()
time.sleep(5)
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
hd = []
for i in range(251) :
    r = soup.select('tbody')[5]
    r = r.select('tr')[i].contents[1].text.replace("펼치기","").strip()
    hd.append(r)
headers = []
for i in range(11,18):
    if i != 16 and i != 17:
        ext=str(soup.select('th')[i].select('div')[0])
        idx=ext.find(">")
        ext=ext[idx+1:].lstrip()
        headers.append(ext[:7])
    else:
        ext=str(soup.select('th')[i].select('div')[0])
        idx=ext.find(">")
        ext=ext[idx+1:].lstrip()
        if i == 16:
            headers.append(ext[:10])
        else:
            headers.append(ext[:4])
print(headers)
df = {'주요재무정보': hd}
df = pd.DataFrame(df)
for j in range(0,len(headers)) :
    fs = []
    for i in range(251) :
        r = soup.select('tbody')[5]
        r = r.select('tr')[i] #행 0~
        r = r.select('td')[j+1].text
        fs.append(r)
    df[headers[j]] = fs
df.to_excel(writer,sheet_name="Jeamu_state",float_format='%30.1f')
driver.close()
#%%
# 재무분석 - 현금흐름표 ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
import time

url = f'https://navercomp.wisereport.co.kr/v2/company/c1030001.aspx?cmp_cd={ticker}'
driver = webdriver.Chrome('./chromedriver')
driver.get(url)
time.sleep(5)
driver.find_element_by_id('rpt_td3').click()
time.sleep(5)
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
hd = []
r = soup.select('tbody')[5]
print(len(r.select('tr')))
for i in range(312) :
    try:
        r = soup.select('tbody')[5]
        r = r.select('tr')[i].contents[1].text.replace("펼치기","").strip()
        hd.append(r)
    except:
        print(f"{ticker} Error Find")
        break
headers = []

for i in range(11,18):
    if i != 16 and i != 17:
        ext=str(soup.select('th')[i].select('div')[0])
        idx=ext.find(">")
        ext=ext[idx+1:].lstrip()
        headers.append(ext[:7])
    else:
        ext=str(soup.select('th')[i].select('div')[0])
        idx=ext.find(">")
        ext=ext[idx+1:].lstrip()
        if i == 16:
            headers.append(ext[:10])
        else:
            headers.append(ext[:4])
print(headers)
df = {'주요재무정보': hd}
df = pd.DataFrame(df)
for j in range(0,len(headers)) :
    fs = []
    for i in range(312) :
        try:
            r = soup.select('tbody')[5]
            r = r.select('tr')[i] #행 0~
            r = r.select('td')[j+1].text
            fs.append(r)
        except:
            print(f"{ticker} Error Find")
            break
    df[headers[j]] = fs
df.to_excel(writer,sheet_name="heon_Flow",float_format='%30.1f')
driver.close()

#%%
# 투자지표 - 수익성 ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
url = f'https://navercomp.wisereport.co.kr/v2/company/c1040001.aspx?cmp_cd={ticker}'
driver = webdriver.Chrome('./chromedriver')
driver.get(url)
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
hd = []
for i in range(0,46,2) :
    r = soup.select('tbody')[6]
    r = r.select('span')[i].contents[2].strip()
    hd.append(r)
headers = []
for i in range(14,21):
    ext = str(soup('th')[i]).split("\">")[2]
    idx=ext.find("<br/>")
    headers.append(ext[:idx])
print(headers)
df = {'주요재무정보': hd}
df = pd.DataFrame(df)
for j in range(1,len(headers)) :
    fs = []
    for i in range(0,23) :
        r = soup.select('tbody')[6]
        r = r.select('tr')[i]
        r = r.select('td')[j].text
        fs.append(r)
    df[headers[j]] = fs
df.to_excel(writer,sheet_name="su_ik",float_format='%30.1f')
driver.close()
#%%
# 재무분석 - 안정성 ■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd
url = f'https://navercomp.wisereport.co.kr/v2/company/c1040001.aspx?cmp_cd={ticker}'
driver = webdriver.Chrome('./chromedriver')
driver.get(url)
driver.find_element_by_id('val_tab3').click()
html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
hd = []
for i in range(0,56,2) :
    r = soup.select('tbody')[6]
    r = r.select('span')[i].contents[2].strip()
    hd.append(r)
headers = []
for i in range(14,21):
    ext = str(soup('th')[i]).split("\">")[2]
    idx=ext.find("<br/>")
    headers.append(ext[:idx])
print(headers)
df = {'주요재무정보': hd}
df = pd.DataFrame(df)
for j in range(1,len(headers)) :
    fs = []
    for i in range(0,28) :
        r = soup.select('tbody')[6]
        r = r.select('tr')[i]
        r = r.select('td')[j].text
        fs.append(r)
    df[headers[j]] = fs
df.to_excel(writer,sheet_name="an_jung",float_format='%30.1f')
driver.close()

#%%
#%%
from bs4 import BeautifulSoup
from selenium import webdriver
import pandas as pd

url = 'https://navercomp.wisereport.co.kr/v2/company/c1010001.aspx?cmp_cd=005930'

driver = webdriver.Chrome()
driver.get(url)

html = driver.page_source
soup = BeautifulSoup(html, 'html.parser')
r = soup.select('dt')
content= r[11].text.replace(" ","").split(":")
content+=r[12].text.split(":")
content.pop(2)
content.append(ticker)
df=pd.DataFrame(content)
df.to_excel(writer,sheet_name="jongmock",float_format='%30.1f')
driver.close()
writer.save()